from openpyxl import *
from tkinter import *

wb = load_workbook('sheet.xlsx')
sheet = wb.active

def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    sheet.cell(row=1, column=1).value = 'Department'
    sheet.cell(row=1, column=2).value = 'School'
    sheet.cell(row=1, column=3).value = 'Number'

def focus1(event):
    department_field.focus_set()

def focus2(event):
    school_field.focus_set()

def focus3(event):
    number_field.focus_set()


def clear():
    department_field.delete(0, END)
    school_field.delete(0, END)
    number_field.delete(0, END)

def insert():
    if (department_field.get() == "" and
        school_field.get() == "" and
        number_field.get() == ""): 

        print("empty input")

    else:
        current_row  = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = department_field.get()
        sheet.cell(row=current_row + 1, column=2).value = school_field.get()
        sheet.cell(row=current_row + 1, column=3).value = number_field.get()

        wb.save('sheet.xlsx')

        department_field.focus_set()

        clear()

def key(event):
    if (event.char) == '\r':
        insert()

if __name__ == '__main__':
    root = Tk()
    root.configure(background='light blue')
    root.title('tkinter')
    root.geometry('550x300')

    excel()

    heading = Label(root, text='Form', bg='light blue')
    department = Label(root, text='Department', bg='light blue')
    school = Label(root, text='School', bg='light blue')
    number = Label(root, text='Number', bg='light blue')

    heading.grid(row=0, column=1)
    department.grid(row=1, column=0)
    school.grid(row=2, column=0)
    number.grid(row=3, column=0)

    department_field = Entry(root)
    school_field = Entry(root)
    number_field = Entry(root)

    department_field.bind('<Return>', focus1)
    school_field.bind('<Return>', focus2)
    number_field.bind('<Return>', focus3)

    department_field.grid(row=1, column=1, ipadx='100')
    school_field.grid(row=2, column=1, ipadx='100')
    number_field.grid(row=3, column=1, ipadx='100')

    excel()

    root.bind('<Key>', key)
    root.grid()

    submit = Button(root, text='Submit', fg='Black', bg='Red', command=insert)
    submit.grid(row=9, column=1)
    

    root.mainloop()